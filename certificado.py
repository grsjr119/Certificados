from openpyxl import Workbook
from openpyxl import load_workbook

# FILE_PATH = 'income.xlsx'
# SHEET = 'Certif Mat'

# workbook = load_workbook(FILE_PATH, read_only=True)
# sheet = workbook[SHEET]

# for row in sheet.iter_rows():
#     print(row[0].value)

# print("se pudo prro")
# hoja2 = workbook.create_sheet("Hoja")

# workbook.create_sheet(index=1, title="Otra hoja")
# print("Ejecusion correcta")

#
wb2 = load_workbook('nuevo.xlsx')



ws1 = wb2.create_sheet("prueba")

wb2.save('nuevo.xlsx')



